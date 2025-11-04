import logging
from typing import List, Dict, Any
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from products.models import Product
from products.services import ProductService


logger = logging.getLogger(__name__)


class ExcelProductImporter:
    REQUIRED_COLUMNS = ['name', 'price', 'stock_quantity']
    OPTIONAL_COLUMNS = ['description']
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'failed': 0,
            'errors': []
        }
    
    def validate_file(self) -> None:
        if not self.file_path.exists():
            raise CommandError(f"File not found: {self.file_path}")
        
        if not self.file_path.suffix.lower() in ['.xlsx', '.xls']:
            raise CommandError(f"Invalid file type. Expected Excel file (.xlsx or .xls)")
    
    def load_workbook_data(self) -> List[Dict[str, Any]]:
        try:
            workbook = load_workbook(self.file_path, data_only=True)
            worksheet = workbook.active
            
            if worksheet is None or worksheet.max_row <= 1:
                raise CommandError("Excel file is empty or has no data")
            
            products_data = self._extract_products_from_worksheet(worksheet)
            return products_data
            
        except Exception as e:
            raise CommandError(f"Failed to load Excel file: {str(e)}")
    
    def _extract_products_from_worksheet(
        self,
        worksheet: Worksheet
    ) -> List[Dict[str, Any]]:
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
        
        self._validate_headers(headers)
        
        products_data = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            product_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    product_dict[header] = row[idx]
            
            if product_dict.get('name'):
                product_dict['_row_number'] = row_idx
                products_data.append(product_dict)
        
        return products_data
    
    def _validate_headers(self, headers: List[str]) -> None:
        missing_columns = set(self.REQUIRED_COLUMNS) - set(headers)
        if missing_columns:
            raise CommandError(
                f"Missing required columns: {', '.join(missing_columns)}. "
                f"Expected columns: {', '.join(self.REQUIRED_COLUMNS)}"
            )
    
    def validate_product_data(self, product_data: Dict[str, Any]) -> Dict[str, Any]:
        validated = {}
        
        name = str(product_data.get('name', '')).strip()
        if not name:
            raise ValueError("Product name cannot be empty")
        if len(name) > 255:
            raise ValueError(f"Product name too long (max 255 characters): {name[:50]}...")
        validated['name'] = name
        
        try:
            price = Decimal(str(product_data.get('price', 0)))
            if price <= 0:
                raise ValueError(f"Price must be positive: {price}")
            validated['price'] = price
        except (InvalidOperation, ValueError, TypeError) as e:
            raise ValueError(f"Invalid price value: {product_data.get('price')}")
        
        try:
            stock = int(product_data.get('stock_quantity', 0))
            if stock < 0:
                raise ValueError(f"Stock quantity cannot be negative: {stock}")
            validated['stock_quantity'] = stock
        except (ValueError, TypeError):
            raise ValueError(f"Invalid stock quantity: {product_data.get('stock_quantity')}")
        
        validated['description'] = str(product_data.get('description', '')).strip()
        
        return validated
    
    def import_products(self, products_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        self.stats['total'] = len(products_data)
        
        for product_data in products_data:
            row_number = product_data.get('_row_number', 'Unknown')
            
            try:
                validated_data = self.validate_product_data(product_data)
                
                product, created = Product.objects.update_or_create(
                    name=validated_data['name'],
                    defaults={
                        'description': validated_data['description'],
                        'price': validated_data['price'],
                        'stock_quantity': validated_data['stock_quantity']
                    }
                )
                
                if created:
                    self.stats['created'] += 1
                    logger.info(f"Row {row_number}: Created product '{product.name}'")
                else:
                    self.stats['updated'] += 1
                    logger.info(f"Row {row_number}: Updated product '{product.name}'")
                    
            except ValueError as e:
                self.stats['failed'] += 1
                error_msg = f"Row {row_number}: {str(e)}"
                self.stats['errors'].append(error_msg)
                logger.error(error_msg)
                
            except Exception as e:
                self.stats['failed'] += 1
                error_msg = f"Row {row_number}: Unexpected error - {str(e)}"
                self.stats['errors'].append(error_msg)
                logger.error(error_msg)
        
        return self.stats


class Command(BaseCommand):
    help = 'Import products from an Excel file'
    
    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='Path to the Excel file containing product data'
        )
        
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Enable verbose output'
        )
    
    def handle(self, *args, **options):
        file_path = options['file_path']
        verbose = options.get('verbose', False)
        
        if verbose:
            logger.setLevel(logging.DEBUG)
        
        self.stdout.write(
            self.style.SUCCESS(f'\n=== Product Import Started ===')
        )
        self.stdout.write(f'File: {file_path}\n')
        
        try:
            importer = ExcelProductImporter(file_path)
            
            importer.validate_file()
            self.stdout.write(self.style.SUCCESS('[OK] File validation passed'))
            
            products_data = importer.load_workbook_data()
            self.stdout.write(
                self.style.SUCCESS(f'[OK] Loaded {len(products_data)} products from Excel')
            )
            
            self.stdout.write('\nImporting products...\n')
            stats = importer.import_products(products_data)
            
            self._display_results(stats)
            
        except CommandError as e:
            self.stdout.write(self.style.ERROR(f'\n[ERROR] {str(e)}'))
            raise
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'\n[ERROR] Unexpected error: {str(e)}')
            )
            raise CommandError(str(e))
    
    def _display_results(self, stats: Dict[str, Any]) -> None:
        self.stdout.write('\n' + '=' * 50)
        self.stdout.write(self.style.SUCCESS('\n=== Import Summary ===\n'))
        
        self.stdout.write(f"Total rows processed: {stats['total']}")
        self.stdout.write(
            self.style.SUCCESS(f"[+] Created: {stats['created']}")
        )
        self.stdout.write(
            self.style.WARNING(f"[~] Updated: {stats['updated']}")
        )
        
        if stats['failed'] > 0:
            self.stdout.write(
                self.style.ERROR(f"[-] Failed: {stats['failed']}")
            )
            
            if stats['errors']:
                self.stdout.write('\n' + self.style.ERROR('Errors:'))
                for error in stats['errors'][:10]:
                    self.stdout.write(f"  - {error}")
                
                if len(stats['errors']) > 10:
                    self.stdout.write(
                        f"  ... and {len(stats['errors']) - 10} more errors"
                    )
        
        self.stdout.write('\n' + '=' * 50 + '\n')
        
        if stats['failed'] == 0:
            self.stdout.write(
                self.style.SUCCESS('[SUCCESS] Import completed successfully!\n')
            )
        else:
            self.stdout.write(
                self.style.WARNING('[WARNING] Import completed with errors.\n')
            )

